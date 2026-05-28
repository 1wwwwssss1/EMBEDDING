import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()

# ===== 样式定义 =====
thin = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
header_font = Font(bold=True, size=11)
body_font = Font(size=10)
wrap = Alignment(wrap_text=True, vertical='top')

def write_header(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def write_row(ws, row, data):
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = body_font
        cell.alignment = wrap
        cell.border = thin

# ========== Sheet 1: 所有指标 ==========
ws1 = wb.active
ws1.title = "所有指标"
write_header(ws1, ["序号", "指标名称", "计算规则", "来源与设定"], row=1)

data = [
    (1, "日均使用人数 (avg_dau)", "每日活跃用户数（dau）的算术平均值，向上取整",
     "对每天的去除重复 uid 计数后，除以总天数（对 daily_stats 列表求 dau 字段的平均值，再 math.ceil）"),

    (2, "累计试用人数 (total_trial)", "整个时间范围内所有去重的 uid 总数",
     "对 DataFrame df 的 uid 列取 nunique() 得出"),

    (3, "功能调用总量 (total_calls)", "去重后的指纹数量",
     "指纹 fingerprint = {完整时间戳}_{uid}_{业务类型}，保证同一次调用唯一；df 已对 fingerprint 去重，取 len(df)"),

    (4, "日均活跃率 (avg_active_rate)", "每日活跃率（dau / 累计人数）的平均值，保留一位小数百分比",
     "每日活跃率 = dau / cum，对所有天的活跃率求平均后乘以 100 并格式化"),

    (5, "今日成功转人工次数 (today_handoff_count)", "当天业务类型为'转人工'的记录条数",
     "从 df 中过滤 day == today_str 且 biz == '转人工'，取行数"),

    (6, "日活跃用户 (dau)", "当天 day_full == 该日期的去重 uid 数量",
     "遍历日期范围内每一天，按 day_full 过滤 df，对 uid 列 nunique()"),

    (7, "当日新增 (new)", "当天首次出现的 uid 数量",
     "通过集合 cum_users 累积所有出现过的 uid，当天 de_uid 减去已有集合 = 新增人数"),

    (8, "累计人数 (cum)", "截至当天累积去重 uid 总数",
     "每过一天将当天 de_uid 并入 cum_users，并取当前集合长度"),

    (9, "部门调用次数 TOP5", "按 (dept_short, mgr) 分组统计指纹行数，取前5",
     "df.groupby(['dept_short','mgr']).size()，降序排列，head(5)"),

    (10, "人员调用次数 TOP15", "按 (name, dept_short, mgr) 分组统计指纹行数，取前15",
     "df.groupby(['name','dept_short','mgr']).size()，降序排列，head(15)"),

    (11, "部门人数", "一级总人数取自 L1_DEPT_CONFIG[l1]['total']；二级人数取自 L2_HEADCOUNT；直属 = 一级总人数 - 已知二级人数之和",
     "配置字典人工设定，用于计算部门使用人数占比"),

    (12, "调用功能人数", "每个二级部门（或直属）的去重 uid 数量",
     "按部门过滤 df 后对 uid 列 nunique()"),

    (13, "部门使用人数占比", "调用功能人数 / 部门人数 × 100%（保留一位小数）",
     "直接在生成表格时计算"),

    (14, "每日功能调用次数（询价/面单/轨迹）", "当天该业务类型的指纹数量",
     "按 day 和 biz 过滤 df，取行数"),

    (15, "各功能模块累计使用人数", "该业务类型下所有去重 uid 数",
     "s_df['uid'].nunique()"),

    (16, "各功能模块调用次数", "该业务类型下的指纹总数",
     "len(s_df)"),

    (17, "各功能模块占比", "调用次数 / 总调用次数 × 100%",
     "c_count / total_calls * 100"),

    (18, "部门 TOP5（特定功能）", "按 (dept_short, mgr) 分组，统计 uid 去重数和行数，取前5",
     "s_df.groupby(['dept_short','mgr']).agg(uid=('uid','nunique'), cnt=('biz','count')).sort_values('cnt',ascending=False).head(5)"),

    (19, "人员 TOP10（特定功能）", "按 (name, mgr) 分组统计行数，取前10",
     "s_df.groupby(['name','mgr']).size().sort_values(ascending=False).head(10)"),

    (20, "询价用户数 / 工单用户数（每日）", "当天该功能类型下去重 uid 数量",
     "按 day 和 biz 过滤 df，取 uid nunique()"),

    (21, "部门层级来源", "一级部门 l1 = dept.split('/')[1]；二级部门 l2 = dept.split('/')[2]",
     "CSV 或 JSONL 中的 department 字段使用斜杠分隔，取第2层为一级，第3层为二级"),

    (22,
     "【agent.log】询价识别规则",
     "满足以下任一条件即标记 biz='询价'：\n"
     "① 行包含 \"agent.tools.quote_tool\" 且包含 \"request(async)\"\n"
     "② 行包含 \"quote_tool\" 且包含 \"tool_calls=[\"",
     "来源：agent.log 主循环\n"
     "方式：Python in 字符串包含判断\n"
     "优先级：最高（询价判断优先于轨迹查询/面单推送/转人工）\n\n"
     "匹配示例1：2026-04-25 10:30:00 [INFO] agent.tools.quote_tool request(async) ...\n"
     "匹配示例2：2026-04-25 10:30:00 [INFO] tool_calls=[{\"name\":\"quote_tool\",\"args\":{...}}]"),

    (23,
     "【agent.log】轨迹查询识别规则",
     "满足以下任一条件即标记 biz='轨迹查询'：\n"
     "① 行包含 \"api.track.service\" 且包含 \"地理编码线程已启动\"\n"
     "② 行包含 \"track\" 且包含 \"tool_calls=[\"",
     "来源：agent.log 主循环\n"
     "方式：Python in 字符串包含判断\n"
     "优先级：次于询价，高于面单推送和转人工\n\n"
     "匹配示例1：2026-04-25 11:00:00 [INFO] api.track.service 地理编码线程已启动\n"
     "匹配示例2：2026-04-25 11:00:00 [INFO] tool_calls=[{\"name\":\"track_tool\",\"args\":{...}}]"),

    (24,
     "【agent.log】面单推送识别规则",
     "满足以下条件即标记 biz='面单推送'：\n"
     "行包含 \"waybill\" 且包含 \"tool_calls=[\"",
     "来源：agent.log 主循环\n"
     "方式：Python in 字符串包含判断\n"
     "优先级：次于询价和轨迹查询，高于转人工\n\n"
     "匹配示例：2026-04-25 11:15:00 [INFO] tool_calls=[{\"name\":\"waybill_tool\",\"args\":{...}}]"),

    (25,
     "【agent.log】转人工识别规则（旧格式：日期 < 2026-04-26）",
     "满足以下条件即立即标记 biz='转人工'：\n"
     "行包含 \"intents=[\" 且包含 \"'转人工'\"",
     "来源：agent.log 主循环\n"
     "方式：Python in 字符串包含判断\n"
     "适用日期范围：log_dt < datetime(2026, 4, 26)\n\n"
     "匹配示例：2026-04-25 12:00:00 [INFO] intents=[\"转人工\"]\n"
     "结果：直接写入 raw_rows，biz='转人工'"),

    (26,
     "【agent.log】转人工识别规则（新格式：日期 >= 2026-04-26）",
     "分两阶段：\n"
     "① 发现行包含 \"intents=[\" 且 \"'转人工'\" 时，不立即记录，而是设置 pending_handoff[current_uid] = True 并跳过本次循环\n"
     "② 后续同一 uid 的行中，出现 \"tracking_tool\" 或 \"api.track.service\" 时，才标记 biz='转人工'，并重置 pending_handoff[current_uid] = False",
     "来源：agent.log 主循环\n"
     "适用日期范围：log_dt >= datetime(2026, 4, 26)\n"
     "设计意图：4月26日后，转人工不再由 intent 直接标记，而是转入人工后用户实际使用了轨迹查询功能才视为一次成功的转人工\n\n"
     "匹配示例：\n"
     "  第1行：2026-04-26 12:00:00 [INFO] intents=[\"转人工\"] → 设置 pending_handoff[uid]=True\n"
     "  后续行：2026-04-26 12:00:05 [INFO] tool_calls=[{\"name\":\"tracking_tool\",...}] → biz='转人工'"),

    (27,
     "【conversations/audit JSONL】询价识别规则",
     "判断前提：从 obj.get(\"tool_calls\") 中提取 tool_names 列表\n"
     "匹配条件：any(\"quote\" in n for n in tool_names)\n"
     "匹配到则 biz = \"询价\"",
     "来源：parse_jsonl_conversations() 函数\n"
     "匹配方式：遍历 tool_calls 中每个对象的 tool_name 或 name 字段，包含 \"quote\" 即命中\n"
     "优先级：最高（询价判断优先于轨迹查询/面单推送/转人工）\n\n"
     "匹配示例：tool_calls[0] = {\"tool_name\":\"quote_tool\"} 或 {\"name\":\"quote_price\"}"),

    (28,
     "【conversations/audit JSONL】轨迹查询识别规则",
     "判断前提：从 obj.get(\"tool_calls\") 中提取 tool_names 列表\n"
     "匹配条件：any(\"track\" in n for n in tool_names)\n"
     "匹配到则 biz = \"轨迹查询\"",
     "来源：parse_jsonl_conversations() 函数\n"
     "匹配方式：遍历 tool_calls 中每个对象的 tool_name 或 name 字段，包含 \"track\" 即命中\n"
     "优先级：次于询价，高于面单推送和转人工\n\n"
     "匹配示例：tool_calls[0] = {\"tool_name\":\"tracking_tool\"} 或 {\"name\":\"track_order\"}"),

    (29,
     "【conversations/audit JSONL】面单推送识别规则",
     "判断前提：从 obj.get(\"tool_calls\") 中提取 tool_names 列表\n"
     "匹配条件：any(\"waybill\" in n for n in tool_names)\n"
     "匹配到则 biz = \"面单推送\"",
     "来源：parse_jsonl_conversations() 函数\n"
     "匹配方式：遍历 tool_calls 中每个对象的 tool_name 或 name 字段，包含 \"waybill\" 即命中\n"
     "优先级：次于询价和轨迹查询，高于转人工\n\n"
     "匹配示例：tool_calls[0] = {\"tool_name\":\"waybill_tool\"} 或 {\"name\":\"print_waybill\"}"),

    (30,
     "【conversations/audit JSONL】转人工识别规则",
     "匹配条件：\"转人工\" in str(obj.get(\"intent\") or \"\")\n"
     "匹配到则 biz = \"转人工\"",
     "来源：parse_jsonl_conversations() 函数\n"
     "匹配方式：提取 JSONL 中 intent 字段，检查是否包含字符串\"转人工\"\n"
     "优先级：最低——仅当上述 询价/轨迹/面单 均未匹配时才判断 intent\n\n"
     "匹配示例：\"intent\": \"转人工\" 或 \"intent\": \"用户要求转人工\"\n"
     "注意：不区分日期，所有日期统一通过 intent 判断"),

    (31,
     "【conversations/audit JSONL】业务判断优先级",
     "严格按以下顺序判断（if-elif-elif-elif 链）：\n"
     "① quote → 询价\n"
     "② track → 轨迹查询\n"
     "③ waybill → 面单推送\n"
     "④ 转人工 → 转人工\n\n"
     "如果同时匹配多个，只取第一个匹配到的业务类型",
     "来源：parse_jsonl_conversations() 函数中的 if-elif-elif-elif 链\n"
     "设计意图：与 agent.log 的识别逻辑保持一致性\n"
     "agent.log 主循环中也是先判断工具调用（询价/轨迹/面单），最后才判断 intent")
]

for i, row_data in enumerate(data, 2):
    write_row(ws1, i, row_data)

# 设置列宽
ws1.column_dimensions['A'].width = 6
ws1.column_dimensions['B'].width = 45
ws1.column_dimensions['C'].width = 75
ws1.column_dimensions['D'].width = 80
# ========== Sheet 2: 表格结构说明 ==========
ws2 = wb.create_sheet("表格结构说明")
write_header(ws2, ["所属报告章节", "表格标题（Excel Sheet名）", "包含字段（列）", "简要说明"], row=1)

table_data = [
    ("一、核心数据概览", "核心数据概览", "指标｜数值", "5个核心KPI的简要数值"),
    ("每日活跃统计", "每日活跃统计", "日期｜日活跃用户｜当日新增｜累计人数", "每日DAU、新增、累积用户变化"),
    ("部门调用 TOP5", "部门调用TOP5", "排名｜部门｜部门负责人｜次数", "所有部门按调用次数排名前5"),
    ("人员调用 TOP15", "人员调用TOP15", "排名｜姓名｜部门｜部门负责人｜调用次数", "所有用户按调用次数排名前15"),
    ("各部门使用情况明细", "各部门使用明细", "一级部门｜负责人｜二级部门｜部门人数｜调用功能人数｜部门使用人数占比", "按一级部门分组，每个二级部门一行，含覆盖率"),
    ("每日功能调用统计", "每日功能调用统计", "日期｜询价调用次数｜面单推送调用次数｜轨迹查询调用次数", "三大核心功能的每日调用量"),
    ("功能应用详情（各功能）", "功能-询价 / 功能-面单推送 / 功能-轨迹查询 / 功能-转人工", "部门TOP5表：部门｜负责人｜人数｜次数\n人员TOP10表：姓名｜负责人｜次数", "每个功能分别分析部门分布和个人高频用户"),
    ("重点功能渗透情况", "重点功能渗透情况", "日期｜询价用户数｜工单(转人工)用户数", "每日使用询价和转人工的用户数，观察渗透趋势"),
]

for i, row_data in enumerate(table_data, 2):
    write_row(ws2, i, row_data)

ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 45
ws2.column_dimensions['C'].width = 60
ws2.column_dimensions['D'].width = 55

# ========== Sheet 3: 业务识别流程图 ==========
ws3 = wb.create_sheet("业务识别流程图")

write_header(ws3, ["步骤", "数据来源", "描述/判断条件", "结果/说明"], row=1)

flow_data = [
    ("1", "agent.log / conversations", "获取数据源，提取 uid、时间戳、行/对象内容", "待处理的原始记录"),
    ("2", "CSV / JSONL 内嵌", "提取用户信息：name / dept / mgr", "user_full_map[uid] 或 fallback 兜底"),
    ("3a", "agent.log", "匹配询价：行含 'agent.tools.quote_tool' + 'request(async)' 或 'quote_tool' + 'tool_calls=['", "biz = '询价'"),
    ("3b", "agent.log", "匹配轨迹查询：行含 'api.track.service' + '地理编码线程已启动' 或 'track' + 'tool_calls=['", "biz = '轨迹查询'"),
    ("3c", "agent.log", "匹配面单推送：行含 'waybill' + 'tool_calls=['", "biz = '面单推送'"),
    ("3d1", "agent.log (旧格式)", "日期 < 2026-04-26：行含 'intents=[' + ''转人工'' → 直接标记", "biz = '转人工'"),
    ("3d2", "agent.log (新格式)", "日期 >= 2026-04-26：先设 pending_handoff=True，后续出现 'tracking_tool' 或 'api.track.service' 才标记", "biz = '转人工'"),
    ("4a", "conversations JSONL", "匹配询价：tool_names 中含 'quote'", "biz = '询价'"),
    ("4b", "conversations JSONL", "匹配轨迹查询：tool_names 中含 'track'", "biz = '轨迹查询'"),
    ("4c", "conversations JSONL", "匹配面单推送：tool_names 中含 'waybill'", "biz = '面单推送'"),
    ("4d", "conversations JSONL", "匹配转人工：intent 字段含 '转人工'（不区分日期）", "biz = '转人工'"),
    ("5", "所有来源", "按优先级取第一个匹配的业务：询价 > 轨迹查询 > 面单推送 > 转人工", "最终 biz 值"),
    ("6", "所有来源", "生成指纹 fingerprint = {完整时间戳}_{uid}_{biz}，后续去重", "去重后的统计记录"),
]

for i, row_data in enumerate(flow_data, 2):
    write_row(ws3, i, row_data)

ws3.column_dimensions['A'].width = 8
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 80
ws3.column_dimensions['D'].width = 30

# ========== 保存文件 ==========
output_path = "指标分析说明.xlsx"
wb.save(output_path)
print(f"✅ 已生成：{output_path}")